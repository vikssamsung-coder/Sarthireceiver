#!/usr/bin/env python3
"""
Bigul · Sarthi CDP · New Client 360 Extract
============================================

Creates one client-level Excel report for accounts opened in the current
calendar month and the previous calendar month.

Sources
-------
MySQL / sarthi_cdp
  - client_master
  - client_lead_contact_map
  - funds_client_day_fact
  - clientwise_datewise_brokerage
  - margin_sheet
  - orders_client_day_symbol_fact
  - client_subscription_master
  - subscription_master

Files
  - Leads.csv: lead number and campaign attribution
  - Optional TPP workbook/CSV: Date, Client Code, Amount, Product

Examples
--------
python sarthi_new_clients_360_extract.py
python sarthi_new_clients_360_extract.py --leads "D:\\Sarthi\\Leads.csv"
python sarthi_new_clients_360_extract.py --tpp "D:\\Sarthi\\TPP Subscription.xlsx"
python sarthi_new_clients_360_extract.py --window rolling --days 60

Database configuration
----------------------
The script first tries common_config.DB_CONFIG. If unavailable, it uses:
SARTHI_DB_HOST, SARTHI_DB_PORT, SARTHI_DB_NAME, SARTHI_DB_USER,
SARTHI_DB_PASSWORD. A missing password is requested securely at runtime.
"""

from __future__ import annotations

import argparse
import getpass
import os
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable

import pandas as pd
import pymysql


# Sarthi CDP login (hardcoded defaults; env vars still override if set).
DB_HOST = "localhost"
DB_PORT = 3306
DB_USER = "sarthi_user"
DB_PASSWORD = ""
DB_NAME = "sarthi_cdp"

DEFAULT_DB = {
    "host": os.getenv("SARTHI_DB_HOST", DB_HOST),
    "port": int(os.getenv("SARTHI_DB_PORT", str(DB_PORT))),
    "database": os.getenv("SARTHI_DB_NAME", DB_NAME),
    "user": os.getenv("SARTHI_DB_USER", DB_USER),
    "password": os.getenv("SARTHI_DB_PASSWORD", DB_PASSWORD),
    "charset": "utf8mb4",
    "cursorclass": pymysql.cursors.DictCursor,
}

FIXED_LEADS_FILE = Path(r"D:\Sarthi\Leads\Leads.csv")
DEFAULT_LEADS_CANDIDATES = [FIXED_LEADS_FILE]

DEFAULT_TPP_CANDIDATES = [
    Path(r"D:\Customer Final Evaluation\01_Input\Sarthi_360_Source\TPP Subscription.xlsx"),
    Path(r"C:\Users\Vikrant.Dale\Downloads\TPP SUBSCRIPTION Feb23 to Jun26.xlsx"),
    Path(r"D:\Sarthi\TPP SUBSCRIPTION.xlsx"),
]

OUTPUT_COLUMNS = [
    "Client Code",
    "Client Name",
    "Opening Date",
    "Account Age Days",
    "Account Status",
    "Branch Code",
    "Introducer Code",
    "Lead Number",
    "Lead Match Method",
    "Lead Source",
    "Source Campaign",
    "Campaign Name",
    "Campaign ID",
    "Source Medium",
    "Source Content",
    "Source Term",
    "Source URL",
    "Lead Type",
    "Lead Stage",
    "Appsflyer Media Source",
    "Appsflyer Campaign",
    "Appsflyer Campaign ID",
    "Funds Collected",
    "Funds Paid Out",
    "Net Funds",
    "First Fund Date",
    "Last Fund Date",
    "Gross Brokerage",
    "Brokerage MTD",
    "Brokerage Last 30D",
    "First Trade Date",
    "Last Trade Date",
    "Trading Days",
    "Current Margin Date",
    "Current Total Cash",
    "Current Stock",
    "Current Collateral",
    "Current Total Margin",
    "Current Tradeable Margin",
    "Margin Snapshot Available",
    "Top Symbols",
    "Executed Orders",
    "Traded Value",
    "Subscription Purchased",
    "Subscription Purchase Count",
    "Subscription Amount",
    "Subscription Plans",
    "First Subscription Date",
    "Last Subscription Date",
    "TPP Purchased",
    "TPP Purchase Count",
    "TPP Amount",
    "TPP Products",
    "First TPP Date",
    "Last TPP Date",
    "Total Revenue",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract a client-level 360 report for newly opened clients."
    )
    parser.add_argument("--leads", type=Path, help="Path to Leads.csv.")
    parser.add_argument("--tpp", type=Path, help="Optional TPP Excel/CSV path.")
    parser.add_argument(
        "--output",
        type=Path,
        help="Output .xlsx path. Defaults to the current directory.",
    )
    parser.add_argument(
        "--as-of",
        type=lambda x: datetime.strptime(x, "%Y-%m-%d").date(),
        help="Report anchor date in YYYY-MM-DD. Default: latest source date.",
    )
    parser.add_argument(
        "--window",
        choices=["calendar", "rolling"],
        default="calendar",
        help="'calendar' = current and previous calendar month; 'rolling' = N days.",
    )
    parser.add_argument(
        "--days",
        type=int,
        default=60,
        help="Number of days for --window rolling. Default: 60.",
    )
    return parser.parse_args()


def norm_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def norm_code(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip().upper()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def clean_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def clean_id(value: object) -> str:
    """Like clean_text, but strips the trailing '.0' float artifact from
    integer-like IDs (e.g. '1092191.0' -> '1092191'). Preserves case so
    alphanumeric campaign IDs are left intact."""
    text = clean_text(value)
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def first_existing(columns: Iterable[str], aliases: Iterable[str]) -> str | None:
    lookup = {norm_header(col): col for col in columns}
    for alias in aliases:
        hit = lookup.get(norm_header(alias))
        if hit is not None:
            return hit
    return None


def resolve_path(
    supplied: Path | None,
    candidates: list[Path],
    glob_patterns: list[str],
    required: bool,
) -> Path | None:
    if supplied:
        if supplied.exists():
            return supplied
        raise FileNotFoundError(f"File not found: {supplied}")

    for candidate in candidates:
        if candidate.exists():
            return candidate

    search_roots = [
        Path.cwd(),
        Path.home() / "Downloads",
        Path(r"C:\Users\Vikrant.Dale\Downloads"),
        Path(r"D:\Sarthi"),
        Path(r"D:\Sarthi\Leads"),
    ]
    matches: list[Path] = []
    for root in search_roots:
        if not root.exists():
            continue
        for pattern in glob_patterns:
            matches.extend(p for p in root.glob(pattern) if p.is_file())
    if matches:
        return max(matches, key=lambda p: p.stat().st_mtime)

    if required:
        raise FileNotFoundError(
            f"Required Leads.csv was not found at its fixed location: {FIXED_LEADS_FILE}"
        )
    return None


def get_db_config() -> dict:
    try:
        from common_config import DB_CONFIG  # type: ignore

        config = dict(DB_CONFIG)
        config.setdefault("charset", "utf8mb4")
        config.setdefault("cursorclass", pymysql.cursors.DictCursor)
        return config
    except (ImportError, AttributeError):
        config = dict(DEFAULT_DB)
        if not config["password"]:
            if not sys.stdin.isatty():
                raise RuntimeError(
                    "Database password missing. Set SARTHI_DB_PASSWORD."
                )
            config["password"] = getpass.getpass("Sarthi MySQL password: ")
        return config


def read_sql(connection, sql: str, params: tuple = ()) -> pd.DataFrame:
    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        columns = [item[0] for item in cursor.description] if cursor.description else []
    return pd.DataFrame(rows, columns=columns)


def get_source_anchor(connection) -> date:
    sql = """
        SELECT MAX(source_date) AS as_of_date
        FROM (
            SELECT MAX(opening_date) AS source_date FROM client_master
            UNION ALL
            SELECT MAX(txn_date) FROM funds_client_day_fact
            UNION ALL
            SELECT MAX(trade_date) FROM clientwise_datewise_brokerage
            UNION ALL
            SELECT MAX(report_date) FROM margin_sheet
            UNION ALL
            SELECT MAX(txn_date) FROM orders_client_day_symbol_fact
        ) x
    """
    df = read_sql(connection, sql)
    value = df.iloc[0]["as_of_date"]
    if pd.isna(value):
        return date.today()
    return pd.Timestamp(value).date()


def calculate_window(as_of: date, mode: str, days: int) -> tuple[date, date]:
    if mode == "rolling":
        if days < 1:
            raise ValueError("--days must be at least 1.")
        return as_of - timedelta(days=days - 1), as_of

    first_this_month = as_of.replace(day=1)
    previous_month_end = first_this_month - timedelta(days=1)
    return previous_month_end.replace(day=1), as_of


def extract_base_clients(connection, start_date: date, end_date: date) -> pd.DataFrame:
    sql = """
        SELECT
            cm.client_code,
            cm.client_name,
            cm.opening_date,
            cm.status AS account_status,
            cm.branch_code,
            cm.introducer_code,
            clm.matched_lead_code
        FROM client_master cm
        LEFT JOIN client_lead_contact_map clm
          ON clm.client_code COLLATE utf8mb4_unicode_ci =
             cm.client_code COLLATE utf8mb4_unicode_ci
        WHERE cm.opening_date BETWEEN %s AND %s
          AND cm.client_code IS NOT NULL
          AND TRIM(cm.client_code) <> ''
        ORDER BY cm.opening_date, cm.client_code
    """
    df = read_sql(connection, sql, (start_date, end_date))
    if df.empty:
        return df
    df["client_code"] = df["client_code"].map(norm_code)
    df["matched_lead_code"] = df["matched_lead_code"].map(norm_code)
    df = (
        df.sort_values(["opening_date", "client_code"])
        .drop_duplicates("client_code", keep="last")
        .reset_index(drop=True)
    )
    return df


def base_cte() -> str:
    return """
        WITH base AS (
            SELECT DISTINCT client_code, opening_date
            FROM client_master
            WHERE opening_date BETWEEN %s AND %s
              AND client_code IS NOT NULL
              AND TRIM(client_code) <> ''
        )
    """


def extract_funds(connection, start_date: date, end_date: date, as_of: date) -> pd.DataFrame:
    sql = base_cte() + """
        SELECT
            b.client_code,
            COALESCE(SUM(f.receipt_amt_day), 0) AS funds_collected,
            COALESCE(SUM(f.payout_amt_day), 0) AS funds_paid_out,
            COALESCE(SUM(f.net_amt_day), 0) AS net_funds,
            MIN(CASE WHEN f.receipt_amt_day > 0 THEN f.txn_date END) AS first_fund_date,
            MAX(CASE WHEN f.receipt_amt_day > 0 THEN f.txn_date END) AS last_fund_date
        FROM base b
        LEFT JOIN funds_client_day_fact f
          ON f.client_code COLLATE utf8mb4_unicode_ci =
             b.client_code COLLATE utf8mb4_unicode_ci
         AND f.txn_date BETWEEN b.opening_date AND %s
        GROUP BY b.client_code
    """
    return read_sql(connection, sql, (start_date, end_date, as_of))


def extract_brokerage(
    connection, start_date: date, end_date: date, as_of: date
) -> pd.DataFrame:
    month_start = as_of.replace(day=1)
    last_30_start = as_of - timedelta(days=29)
    sql = base_cte() + """
        SELECT
            b.client_code,
            COALESCE(SUM(br.gross_brok), 0) AS gross_brokerage,
            COALESCE(SUM(CASE WHEN br.trade_date BETWEEN %s AND %s
                              THEN br.gross_brok ELSE 0 END), 0) AS brokerage_mtd,
            COALESCE(SUM(CASE WHEN br.trade_date BETWEEN %s AND %s
                              THEN br.gross_brok ELSE 0 END), 0) AS brokerage_last_30d,
            MIN(CASE WHEN br.gross_brok > 0 THEN br.trade_date END) AS first_trade_date,
            MAX(CASE WHEN br.gross_brok > 0 THEN br.trade_date END) AS last_trade_date,
            COUNT(DISTINCT CASE WHEN br.gross_brok > 0 THEN br.trade_date END) AS trading_days
        FROM base b
        LEFT JOIN clientwise_datewise_brokerage br
          ON br.client_code COLLATE utf8mb4_unicode_ci =
             b.client_code COLLATE utf8mb4_unicode_ci
         AND br.trade_date BETWEEN b.opening_date AND %s
        GROUP BY b.client_code
    """
    params = (
        start_date,
        end_date,
        month_start,
        as_of,
        last_30_start,
        as_of,
        as_of,
    )
    return read_sql(connection, sql, params)


def extract_margin(
    connection, start_date: date, end_date: date, as_of: date
) -> tuple[pd.DataFrame, date | None]:
    margin_date_sql = "SELECT MAX(report_date) AS margin_date FROM margin_sheet WHERE report_date <= %s"
    date_df = read_sql(connection, margin_date_sql, (as_of,))
    margin_date = date_df.iloc[0]["margin_date"]
    if pd.isna(margin_date):
        return pd.DataFrame(columns=["client_code"]), None
    margin_date = pd.Timestamp(margin_date).date()

    sql = base_cte() + """
        SELECT
            b.client_code,
            m.report_date AS current_margin_date,
            m.total_cash AS current_total_cash,
            m.total_stock AS current_stock,
            m.total_collateral AS current_collateral,
            m.total_margin AS current_total_margin,
            m.tradeable_margin AS current_tradeable_margin
        FROM base b
        LEFT JOIN margin_sheet m
          ON m.client_code COLLATE utf8mb4_unicode_ci =
             b.client_code COLLATE utf8mb4_unicode_ci
         AND m.report_date = %s
    """
    return (
        read_sql(connection, sql, (start_date, end_date, margin_date)),
        margin_date,
    )


def extract_top_symbols(
    connection, start_date: date, end_date: date, as_of: date
) -> pd.DataFrame:
    sql = base_cte() + """
        , symbol_totals AS (
            SELECT
                b.client_code,
                s.trading_symbol,
                SUM(s.executed_orders_symbol_day) AS executed_orders,
                SUM(s.traded_value_symbol_day) AS traded_value
            FROM base b
            JOIN orders_client_day_symbol_fact s
              ON s.client_code COLLATE utf8mb4_unicode_ci =
                 b.client_code COLLATE utf8mb4_unicode_ci
             AND s.txn_date BETWEEN b.opening_date AND %s
            WHERE s.executed_orders_symbol_day > 0
            GROUP BY b.client_code, s.trading_symbol
        ),
        ranked AS (
            SELECT
                client_code,
                trading_symbol,
                executed_orders,
                traded_value,
                ROW_NUMBER() OVER (
                    PARTITION BY client_code
                    ORDER BY executed_orders DESC, traded_value DESC, trading_symbol
                ) AS symbol_rank
            FROM symbol_totals
        ),
        top_symbols AS (
            SELECT
                client_code,
                GROUP_CONCAT(
                    CONCAT(trading_symbol, ' (', executed_orders, ')')
                    ORDER BY symbol_rank SEPARATOR ', '
                ) AS top_symbols
            FROM ranked
            WHERE symbol_rank <= 5
            GROUP BY client_code
        ),
        totals AS (
            SELECT
                client_code,
                SUM(executed_orders) AS executed_orders,
                SUM(traded_value) AS traded_value
            FROM symbol_totals
            GROUP BY client_code
        )
        SELECT
            b.client_code,
            ts.top_symbols,
            COALESCE(t.executed_orders, 0) AS executed_orders,
            COALESCE(t.traded_value, 0) AS traded_value
        FROM base b
        LEFT JOIN top_symbols ts
          ON ts.client_code COLLATE utf8mb4_unicode_ci =
             b.client_code COLLATE utf8mb4_unicode_ci
        LEFT JOIN totals t
          ON t.client_code COLLATE utf8mb4_unicode_ci =
             b.client_code COLLATE utf8mb4_unicode_ci
    """
    return read_sql(connection, sql, (start_date, end_date, as_of))


def extract_subscriptions(
    connection, start_date: date, end_date: date, as_of: date
) -> pd.DataFrame:
    modern_sql = base_cte() + """
        SELECT
            b.client_code,
            csm.purchase_date,
            csm.plan_name,
            csm.amount_paid AS amount
        FROM base b
        JOIN client_subscription_master csm
          ON csm.client_code COLLATE utf8mb4_unicode_ci =
             b.client_code COLLATE utf8mb4_unicode_ci
         AND csm.purchase_date BETWEEN b.opening_date AND %s
    """
    legacy_sql = base_cte() + """
        SELECT
            b.client_code,
            DATE(sm.subscription_date) AS purchase_date,
            sm.plan_name,
            sm.amount
        FROM base b
        JOIN subscription_master sm
          ON sm.client_code COLLATE utf8mb4_unicode_ci =
             b.client_code COLLATE utf8mb4_unicode_ci
         AND DATE(sm.subscription_date) BETWEEN b.opening_date AND %s
    """
    modern = read_sql(connection, modern_sql, (start_date, end_date, as_of))
    legacy = read_sql(connection, legacy_sql, (start_date, end_date, as_of))
    combined = pd.concat([modern, legacy], ignore_index=True)
    if combined.empty:
        return pd.DataFrame(
            columns=[
                "client_code",
                "subscription_purchase_count",
                "subscription_amount",
                "subscription_plans",
                "first_subscription_date",
                "last_subscription_date",
            ]
        )

    combined["client_code"] = combined["client_code"].map(norm_code)
    combined["purchase_date"] = pd.to_datetime(
        combined["purchase_date"], errors="coerce"
    ).dt.date
    combined["plan_name"] = combined["plan_name"].map(clean_text)
    combined["amount"] = pd.to_numeric(combined["amount"], errors="coerce").fillna(0)
    # Both subscription tables can contain the same purchase. This business-key
    # dedupe prevents double counting while retaining purchases unique to either.
    combined = combined.drop_duplicates(
        ["client_code", "purchase_date", "plan_name", "amount"]
    )

    def join_unique(series: pd.Series) -> str:
        values = sorted({clean_text(v) for v in series if clean_text(v)})
        return ", ".join(values)

    return (
        combined.groupby("client_code", as_index=False)
        .agg(
            subscription_purchase_count=("amount", "size"),
            subscription_amount=("amount", "sum"),
            subscription_plans=("plan_name", join_unique),
            first_subscription_date=("purchase_date", "min"),
            last_subscription_date=("purchase_date", "max"),
        )
    )


LEAD_FIELD_ALIASES = {
    "lead_number": ["Lead Number", "Lead No", "Lead Code", "Prospect ID"],
    "client_code": ["Client Code", "LD Client Code", "Terminal Code"],
    "lead_source": ["Lead Source", "Latest Source"],
    "source_campaign": ["Source Campaign", "Latest Campaign"],
    "campaign_name": ["Campaign Name", "CampaignName"],
    "campaign_id": ["Campaign ID", "CampaignID"],
    "source_medium": ["Source Medium", "Latest Medium"],
    "source_content": ["Source Content", "Content"],
    "source_term": ["Source Term", "Term"],
    "source_url": ["Source URL", "Landing Page URL"],
    "lead_type": ["Lead Type"],
    "lead_stage": ["Lead Stage", "Stage"],
    "appsflyer_media_source": [
        "Appsflyer Media Source",
        "AppsFlyer Media Source",
        "Media Source",
    ],
    "appsflyer_campaign": ["Appsflyer Campaign", "AppsFlyer Campaign"],
    "appsflyer_campaign_id": [
        "Appsflyer Campaign ID",
        "AppsFlyer Campaign ID",
    ],
    "modified_on": ["Modified On", "Last Activity Date", "Updated On"],
    "created_on": ["Created On", "Lead Created On", "Prospect Creation Date"],
    "account_opened_date": [
        "LD Account Opened Date",
        "Account Opened Date",
        "Account Opened On",
    ],
}


def read_csv_flexible(path: Path) -> pd.DataFrame:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin1"):
        try:
            return pd.read_csv(path, dtype=str, low_memory=False, encoding=encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    raise RuntimeError(f"Unable to decode {path}: {last_error}")


def load_leads(path: Path) -> pd.DataFrame:
    raw = read_csv_flexible(path)
    mapped: dict[str, pd.Series] = {}
    for target, aliases in LEAD_FIELD_ALIASES.items():
        source = first_existing(raw.columns, aliases)
        mapped[target] = raw[source] if source else pd.Series("", index=raw.index)
    leads = pd.DataFrame(mapped)
    leads["client_code_key"] = leads["client_code"].map(norm_code)
    leads["lead_number_key"] = leads["lead_number"].map(norm_code)
    for field in ("modified_on", "created_on", "account_opened_date"):
        leads[field] = pd.to_datetime(leads[field], errors="coerce")
    leads["_freshness"] = (
        leads["modified_on"]
        .fillna(leads["account_opened_date"])
        .fillna(leads["created_on"])
    )
    leads["_completeness"] = leads[
        [
            "lead_number",
            "lead_source",
            "source_campaign",
            "campaign_name",
            "source_medium",
            "source_content",
        ]
    ].apply(lambda row: sum(bool(clean_text(v)) for v in row), axis=1)
    leads["_row_no"] = range(len(leads))
    return leads


def select_best_lead_match(base: pd.DataFrame, leads: pd.DataFrame) -> pd.DataFrame:
    lead_fields = [key for key in LEAD_FIELD_ALIASES if key not in {
        "client_code", "modified_on", "created_on", "account_opened_date"
    }]
    id_fields = {"lead_number", "campaign_id", "appsflyer_campaign_id"}
    output_rows: list[dict] = []

    by_client = {
        key: frame
        for key, frame in leads[leads["client_code_key"] != ""].groupby(
            "client_code_key", sort=False
        )
    }
    by_lead = {
        key: frame
        for key, frame in leads[leads["lead_number_key"] != ""].groupby(
            "lead_number_key", sort=False
        )
    }

    for row in base.itertuples(index=False):
        client_code = norm_code(row.client_code)
        matched_lead_code = norm_code(row.matched_lead_code)
        candidates = by_client.get(client_code)
        method = "Client Code"
        if candidates is None and matched_lead_code:
            candidates = by_lead.get(matched_lead_code)
            method = "Lead Contact Map"

        record = {"client_code": client_code, "lead_match_method": "Not Matched"}
        if candidates is not None and not candidates.empty:
            best = candidates.sort_values(
                ["_freshness", "_completeness", "_row_no"],
                ascending=[False, False, False],
                na_position="last",
            ).iloc[0]
            record["lead_match_method"] = method
            for field in lead_fields:
                record[field] = (
                    clean_id(best[field])
                    if field in id_fields
                    else clean_text(best[field])
                )
        else:
            for field in lead_fields:
                record[field] = ""
        output_rows.append(record)
    return pd.DataFrame(output_rows)


TPP_FIELD_ALIASES = {
    "tpp_date": ["Date", "Purchase Date", "Transaction Date", "Subscription Date"],
    "client_code": ["Client Code", "ClientCode", "Terminal Code"],
    "amount": ["Amount", "Revenue", "Subscription Amount", "Amount Paid"],
    "product": ["Product", "Product Name", "Plan Name", "Plan"],
}


def load_tpp(path: Path | None) -> pd.DataFrame:
    empty = pd.DataFrame(
        columns=[
            "client_code",
            "tpp_purchase_count",
            "tpp_amount",
            "tpp_products",
            "first_tpp_date",
            "last_tpp_date",
        ]
    )
    if path is None:
        return empty
    if path.suffix.lower() in {".xlsx", ".xls"}:
        raw = pd.read_excel(path, dtype=str)
    else:
        raw = read_csv_flexible(path)

    mapped = {}
    for target, aliases in TPP_FIELD_ALIASES.items():
        source = first_existing(raw.columns, aliases)
        if source is None and target in {"client_code", "amount"}:
            raise ValueError(
                f"TPP file is missing required field '{target}'. "
                f"Available columns: {list(raw.columns)}"
            )
        mapped[target] = raw[source] if source else pd.Series("", index=raw.index)
    tpp = pd.DataFrame(mapped)
    tpp["client_code"] = tpp["client_code"].map(norm_code)
    tpp["amount"] = pd.to_numeric(
        tpp["amount"].astype(str).str.replace(",", "", regex=False),
        errors="coerce",
    ).fillna(0)
    tpp["tpp_date"] = pd.to_datetime(tpp["tpp_date"], errors="coerce").dt.date
    tpp["product"] = tpp["product"].map(clean_text)
    tpp = tpp[tpp["client_code"] != ""].drop_duplicates(
        ["client_code", "tpp_date", "product", "amount"]
    )
    if tpp.empty:
        return empty

    def join_unique(series: pd.Series) -> str:
        return ", ".join(sorted({clean_text(v) for v in series if clean_text(v)}))

    return (
        tpp.groupby("client_code", as_index=False)
        .agg(
            tpp_purchase_count=("amount", "size"),
            tpp_amount=("amount", "sum"),
            tpp_products=("product", join_unique),
            first_tpp_date=("tpp_date", "min"),
            last_tpp_date=("tpp_date", "max"),
        )
    )


def normalize_key(frame: pd.DataFrame) -> pd.DataFrame:
    if "client_code" in frame.columns:
        frame = frame.copy()
        frame["client_code"] = frame["client_code"].map(norm_code)
    return frame


def build_client_detail(
    base: pd.DataFrame,
    lead_matches: pd.DataFrame,
    funds: pd.DataFrame,
    brokerage: pd.DataFrame,
    margin: pd.DataFrame,
    symbols: pd.DataFrame,
    subscriptions: pd.DataFrame,
    tpp: pd.DataFrame,
    as_of: date,
) -> pd.DataFrame:
    result = base.copy()
    for frame in (
        lead_matches,
        funds,
        brokerage,
        margin,
        symbols,
        subscriptions,
        tpp,
    ):
        result = result.merge(normalize_key(frame), on="client_code", how="left")

    numeric_cols = [
        "funds_collected",
        "funds_paid_out",
        "net_funds",
        "gross_brokerage",
        "brokerage_mtd",
        "brokerage_last_30d",
        "trading_days",
        "current_total_cash",
        "current_stock",
        "current_collateral",
        "current_total_margin",
        "current_tradeable_margin",
        "executed_orders",
        "traded_value",
        "subscription_purchase_count",
        "subscription_amount",
        "tpp_purchase_count",
        "tpp_amount",
    ]
    for col in numeric_cols:
        if col not in result:
            result[col] = 0
        result[col] = pd.to_numeric(result[col], errors="coerce").fillna(0)

    text_cols = [
        "top_symbols",
        "subscription_plans",
        "tpp_products",
        "lead_match_method",
        *[
            key
            for key in LEAD_FIELD_ALIASES
            if key not in {
                "client_code",
                "modified_on",
                "created_on",
                "account_opened_date",
            }
        ],
    ]
    for col in text_cols:
        if col not in result:
            result[col] = ""
        result[col] = result[col].fillna("").map(clean_text)

    result["account_age_days"] = (
        pd.Timestamp(as_of) - pd.to_datetime(result["opening_date"])
    ).dt.days
    result["margin_snapshot_available"] = result["current_margin_date"].notna().map(
        {True: "Yes", False: "No"}
    )
    result["subscription_purchased"] = (
        result["subscription_purchase_count"] > 0
    ).map({True: "Yes", False: "No"})
    result["tpp_purchased"] = (result["tpp_purchase_count"] > 0).map(
        {True: "Yes", False: "No"}
    )
    result["total_revenue"] = (
        result["gross_brokerage"]
        + result["subscription_amount"]
        + result["tpp_amount"]
    )

    rename = {
        "client_code": "Client Code",
        "client_name": "Client Name",
        "opening_date": "Opening Date",
        "account_age_days": "Account Age Days",
        "account_status": "Account Status",
        "branch_code": "Branch Code",
        "introducer_code": "Introducer Code",
        "lead_number": "Lead Number",
        "lead_match_method": "Lead Match Method",
        "lead_source": "Lead Source",
        "source_campaign": "Source Campaign",
        "campaign_name": "Campaign Name",
        "campaign_id": "Campaign ID",
        "source_medium": "Source Medium",
        "source_content": "Source Content",
        "source_term": "Source Term",
        "source_url": "Source URL",
        "lead_type": "Lead Type",
        "lead_stage": "Lead Stage",
        "appsflyer_media_source": "Appsflyer Media Source",
        "appsflyer_campaign": "Appsflyer Campaign",
        "appsflyer_campaign_id": "Appsflyer Campaign ID",
        "funds_collected": "Funds Collected",
        "funds_paid_out": "Funds Paid Out",
        "net_funds": "Net Funds",
        "first_fund_date": "First Fund Date",
        "last_fund_date": "Last Fund Date",
        "gross_brokerage": "Gross Brokerage",
        "brokerage_mtd": "Brokerage MTD",
        "brokerage_last_30d": "Brokerage Last 30D",
        "first_trade_date": "First Trade Date",
        "last_trade_date": "Last Trade Date",
        "trading_days": "Trading Days",
        "current_margin_date": "Current Margin Date",
        "current_total_cash": "Current Total Cash",
        "current_stock": "Current Stock",
        "current_collateral": "Current Collateral",
        "current_total_margin": "Current Total Margin",
        "current_tradeable_margin": "Current Tradeable Margin",
        "margin_snapshot_available": "Margin Snapshot Available",
        "top_symbols": "Top Symbols",
        "executed_orders": "Executed Orders",
        "traded_value": "Traded Value",
        "subscription_purchased": "Subscription Purchased",
        "subscription_purchase_count": "Subscription Purchase Count",
        "subscription_amount": "Subscription Amount",
        "subscription_plans": "Subscription Plans",
        "first_subscription_date": "First Subscription Date",
        "last_subscription_date": "Last Subscription Date",
        "tpp_purchased": "TPP Purchased",
        "tpp_purchase_count": "TPP Purchase Count",
        "tpp_amount": "TPP Amount",
        "tpp_products": "TPP Products",
        "first_tpp_date": "First TPP Date",
        "last_tpp_date": "Last TPP Date",
        "total_revenue": "Total Revenue",
    }
    result = result.rename(columns=rename)
    for column in OUTPUT_COLUMNS:
        if column not in result:
            result[column] = ""
    return result[OUTPUT_COLUMNS].sort_values(
        ["Opening Date", "Client Code"], ascending=[False, True]
    )


def build_summary(detail: pd.DataFrame) -> pd.DataFrame:
    total_clients = len(detail)
    rows = [
        ("Accounts Opened", total_clients),
        ("Lead Matched", int((detail["Lead Match Method"] != "Not Matched").sum())),
        ("Lead Match %", (detail["Lead Match Method"] != "Not Matched").mean() if total_clients else 0),
        ("Funded Clients", int((detail["Funds Collected"] > 0).sum())),
        ("Funded %", (detail["Funds Collected"] > 0).mean() if total_clients else 0),
        ("Total Funds Collected", detail["Funds Collected"].sum()),
        ("Current Margin Clients", int((detail["Current Total Margin"] > 0).sum())),
        ("Current Total Margin", detail["Current Total Margin"].sum()),
        ("Traded Clients", int((detail["Gross Brokerage"] > 0).sum())),
        ("Traded %", (detail["Gross Brokerage"] > 0).mean() if total_clients else 0),
        ("Gross Brokerage", detail["Gross Brokerage"].sum()),
        ("Subscription Buyers", int((detail["Subscription Purchase Count"] > 0).sum())),
        ("Subscription Revenue", detail["Subscription Amount"].sum()),
        ("TPP Buyers", int((detail["TPP Purchase Count"] > 0).sum())),
        ("TPP Revenue", detail["TPP Amount"].sum()),
        ("Total Revenue", detail["Total Revenue"].sum()),
    ]
    return pd.DataFrame(rows, columns=["Metric", "Value"])


def build_data_quality(
    detail: pd.DataFrame,
    leads_path: Path,
    tpp_path: Path | None,
    margin_date: date | None,
    start_date: date,
    end_date: date,
    as_of: date,
) -> pd.DataFrame:
    rows = [
        ("Report As Of", as_of),
        ("Account Opening Window Start", start_date),
        ("Account Opening Window End", end_date),
        ("Leads File", str(leads_path)),
        ("TPP File", str(tpp_path) if tpp_path else "Not supplied / not found"),
        ("Latest Global Margin Date", margin_date or "No margin snapshot"),
        ("Client Rows", len(detail)),
        ("Unique Client Codes", detail["Client Code"].nunique()),
        ("Duplicate Client Codes", int(detail["Client Code"].duplicated().sum())),
        ("Lead Not Matched", int((detail["Lead Match Method"] == "Not Matched").sum())),
        (
            "Latest Margin Snapshot Missing",
            int((detail["Margin Snapshot Available"] == "No").sum()),
        ),
        ("TPP Status", "Loaded" if tpp_path else "TPP columns set to zero"),
        (
            "Top Symbol Rule",
            "Top 5 symbols by executed orders; traded value breaks ties",
        ),
        (
            "Subscription Dedupe Rule",
            "Client + purchase date + plan + amount across both subscription tables",
        ),
    ]
    return pd.DataFrame(rows, columns=["Check", "Result"])


def write_excel(
    output_path: Path,
    detail: pd.DataFrame,
    summary: pd.DataFrame,
    quality: pd.DataFrame,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        writer = pd.ExcelWriter(
            output_path,
            engine="xlsxwriter",
            datetime_format="dd-mmm-yyyy",
            date_format="dd-mmm-yyyy",
        )
    except ModuleNotFoundError:
        writer = pd.ExcelWriter(
            output_path,
            engine="openpyxl",
            datetime_format="DD-MMM-YYYY",
            date_format="DD-MMM-YYYY",
        )

    with writer:
        detail.to_excel(writer, sheet_name="Client Detail", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)
        quality.to_excel(writer, sheet_name="Data Quality", index=False)

        if writer.engine == "xlsxwriter":
            workbook = writer.book
            header_fmt = workbook.add_format(
                {
                    "bold": True,
                    "font_color": "white",
                    "bg_color": "#17365D",
                    "border": 1,
                    "align": "center",
                    "valign": "vcenter",
                    "text_wrap": True,
                }
            )
            money_fmt = workbook.add_format({"num_format": "#,##0.00"})
            integer_fmt = workbook.add_format({"num_format": "#,##0"})
            percent_fmt = workbook.add_format({"num_format": "0.0%"})
            date_fmt = workbook.add_format({"num_format": "dd-mmm-yyyy"})

            for sheet_name, frame in [
                ("Client Detail", detail),
                ("Summary", summary),
                ("Data Quality", quality),
            ]:
                ws = writer.sheets[sheet_name]
                ws.freeze_panes(1, 0)
                ws.autofilter(0, 0, len(frame), max(len(frame.columns) - 1, 0))
                ws.set_row(0, 32)
                for col_idx, col in enumerate(frame.columns):
                    ws.write(0, col_idx, col, header_fmt)
                    max_len = max(
                        len(str(col)),
                        min(
                            50,
                            frame[col].fillna("").astype(str).str.len().max()
                            if len(frame)
                            else len(str(col)),
                        ),
                    )
                    ws.set_column(col_idx, col_idx, min(max(max_len + 2, 11), 42))

            detail_ws = writer.sheets["Client Detail"]
            for col in [
                "Funds Collected",
                "Funds Paid Out",
                "Net Funds",
                "Gross Brokerage",
                "Brokerage MTD",
                "Brokerage Last 30D",
                "Current Total Cash",
                "Current Stock",
                "Current Collateral",
                "Current Total Margin",
                "Current Tradeable Margin",
                "Traded Value",
                "Subscription Amount",
                "TPP Amount",
                "Total Revenue",
            ]:
                idx = detail.columns.get_loc(col)
                detail_ws.set_column(idx, idx, 16, money_fmt)
            for col in [
                "Account Age Days",
                "Trading Days",
                "Executed Orders",
                "Subscription Purchase Count",
                "TPP Purchase Count",
            ]:
                idx = detail.columns.get_loc(col)
                detail_ws.set_column(idx, idx, 13, integer_fmt)
            for col in [
                "Opening Date",
                "First Fund Date",
                "Last Fund Date",
                "First Trade Date",
                "Last Trade Date",
                "Current Margin Date",
                "First Subscription Date",
                "Last Subscription Date",
                "First TPP Date",
                "Last TPP Date",
            ]:
                idx = detail.columns.get_loc(col)
                detail_ws.set_column(idx, idx, 14, date_fmt)

            summary_ws = writer.sheets["Summary"]
            for row_idx, metric in enumerate(summary["Metric"], start=1):
                if str(metric).endswith("%"):
                    summary_ws.write_number(
                        row_idx,
                        1,
                        float(summary.iloc[row_idx - 1]["Value"]),
                        percent_fmt,
                    )
                elif any(
                    token in str(metric)
                    for token in ("Funds", "Margin", "Brokerage", "Revenue")
                ):
                    summary_ws.write_number(
                        row_idx,
                        1,
                        float(summary.iloc[row_idx - 1]["Value"]),
                        money_fmt,
                    )
        else:
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter

            for sheet_name, frame in [
                ("Client Detail", detail),
                ("Summary", summary),
                ("Data Quality", quality),
            ]:
                ws = writer.sheets[sheet_name]
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="17365D")
                    cell.alignment = Alignment(
                        horizontal="center", vertical="center", wrap_text=True
                    )
                for idx, col in enumerate(frame.columns, start=1):
                    max_len = max(
                        len(str(col)),
                        min(
                            50,
                            frame[col].fillna("").astype(str).str.len().max()
                            if len(frame)
                            else len(str(col)),
                        ),
                    )
                    ws.column_dimensions[get_column_letter(idx)].width = min(
                        max(max_len + 2, 11), 42
                    )


def main() -> int:
    args = parse_args()
    leads_path = resolve_path(
        args.leads,
        DEFAULT_LEADS_CANDIDATES,
        ["Leads.csv", "leads.csv", "Leads*.csv", "leads*.csv"],
        required=True,
    )
    tpp_path = resolve_path(
        args.tpp,
        DEFAULT_TPP_CANDIDATES,
        ["*TPP*SUBSCRIPTION*.xlsx", "*TPP*.xlsx", "*TPP*.csv"],
        required=False,
    )

    print("=" * 88)
    print("BIGUL · SARTHI CDP · NEW CLIENT 360 EXTRACT")
    print("=" * 88)
    print(f"Leads file : {leads_path}")
    print(f"TPP file   : {tpp_path or 'Not found; TPP columns will be zero'}")

    connection = pymysql.connect(**get_db_config())
    try:
        as_of = args.as_of or get_source_anchor(connection)
        start_date, end_date = calculate_window(
            as_of, args.window, args.days
        )
        print(f"As-of date : {as_of}")
        print(f"AO window  : {start_date} to {end_date}")

        base = extract_base_clients(connection, start_date, end_date)
        if base.empty:
            raise RuntimeError(
                f"No clients found with opening_date from {start_date} to {end_date}."
            )
        print(f"Clients    : {len(base):,}")

        funds = extract_funds(connection, start_date, end_date, as_of)
        brokerage = extract_brokerage(connection, start_date, end_date, as_of)
        margin, margin_date = extract_margin(
            connection, start_date, end_date, as_of
        )
        symbols = extract_top_symbols(connection, start_date, end_date, as_of)
        subscriptions = extract_subscriptions(
            connection, start_date, end_date, as_of
        )
    finally:
        connection.close()

    leads = load_leads(leads_path)
    lead_matches = select_best_lead_match(base, leads)
    tpp = load_tpp(tpp_path)

    detail = build_client_detail(
        base=base,
        lead_matches=lead_matches,
        funds=funds,
        brokerage=brokerage,
        margin=margin,
        symbols=symbols,
        subscriptions=subscriptions,
        tpp=tpp,
        as_of=as_of,
    )
    summary = build_summary(detail)
    quality = build_data_quality(
        detail=detail,
        leads_path=leads_path,
        tpp_path=tpp_path,
        margin_date=margin_date,
        start_date=start_date,
        end_date=end_date,
        as_of=as_of,
    )

    output_path = args.output or Path(r"D:\Customer Final Evaluation\01_Input\Client_360") / (
        f"Sarthi_New_Client_360_{start_date:%Y%m%d}_{end_date:%Y%m%d}.xlsx"
    )
    write_excel(output_path, detail, summary, quality)

    print("-" * 88)
    print(f"Lead matched   : {(detail['Lead Match Method'] != 'Not Matched').sum():,}")
    print(f"Funded clients : {(detail['Funds Collected'] > 0).sum():,}")
    print(f"Traded clients : {(detail['Gross Brokerage'] > 0).sum():,}")
    print(f"Output         : {output_path.resolve()}")
    print("DONE")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"\nERROR: {exc}", file=sys.stderr)
        raise
